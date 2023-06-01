from .models import Student
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from .models import Student, Marks


def filtered_data(request):
    template_name = "studentproject/filterrecord.html"
    filter_marks = request.GET.get('marks')
    filter_name = request.GET.get('name')
    filter_city = request.GET.get('city')
    filter_standard = request.GET.get('standard')

    filtered_student = Student.objects.all()
    if request.method == "GET":
        filters = {}
        if filter_name:
            # filtered_student = Student.objects.filter(
            #     name__icontains=filter_name)
            filters['name__icontains'] = filter_name

        if filter_city:
            filters["city__icontains"] = filter_city
        if filter_marks:
            filters["marks__marks"] = filter_marks
        if filter_standard:
            filters['standard'] = filter_standard
        if filters:
            filtered_student = Student.objects.filter(
                **filters)

    return render(request, template_name, {'student': filtered_student})


def generate_excel(request):
    filter_marks = request.GET.get('marks')
    filter_name = request.GET.get('name')
    filter_city = request.GET.get('city')

    students = Student.objects.all()

    if filter_name:
        students = students.filter(name__icontains=filter_name)

    if filter_city:
        students = students.filter(city__icontains=filter_city)

    # Generate Excel file
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers
    headers = ['Id', 'Name', 'Standard', 'City', 'Marks']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}1'] = header

    # Write student data
    row_num = 2
    for student in students:
        worksheet[f'A{row_num}'] = student.id
        worksheet[f'B{row_num}'] = student.name
        worksheet[f'C{row_num}'] = student.standard

        worksheet[f'D{row_num}'] = student.city
        worksheet[f'E{row_num}'] = student.marks.marks

        row_num += 1

    # Save the workbook
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=filtered_students.xlsx'
    workbook.save(response)

    return response
